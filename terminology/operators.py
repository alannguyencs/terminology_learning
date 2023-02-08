from constants import *
import aljson
from collections import OrderedDict
import openpyxl
import time
from shutil import copyfile
from random import shuffle
import xlsxwriter

def read_input():
    workbook = openpyxl.load_workbook(INPUT_PATH)
    worksheet = workbook.active

    database = OrderedDict() if not os.path.isfile(DATABASE_PATH) else aljson.load(DATABASE_PATH)

    for row in range(2, worksheet.max_row + 1):
        term = str(worksheet.cell(row=row, column=1).value).strip()
        definition = str(worksheet.cell(row=row, column=2).value).strip()
        label = str(worksheet.cell(row=row, column=3).value).strip()
        reference = str(worksheet.cell(row=row, column=4).value).strip()

        database_ = OrderedDict([
            ('definition', definition),
            ('label', label),
            ('reference', reference),
        ])
        database[term] = database_

    if os.path.isfile(DATABASE_PATH):
        time_ = int(time.time())
        backup_path = DATABASE_PATH.replace('database', f'backup/database_{time_}.json')
        copyfile(DATABASE_PATH, backup_path)
    aljson.save(database, DATABASE_PATH)

def list_labels():
    database = aljson.load(DATABASE_PATH)
    labels = []
    for data_ in database.values():
        label = data_['label']
        if label not in labels:
            labels.append(label)
    labels.sort()
    for id, label in enumerate(labels):
        print (f"{id}: {label}")

def write_question(label):
    database = aljson.load(DATABASE_PATH)
    key_candidates = [key for key in database.keys() if database[key]['label']==label]
    shuffle(key_candidates)

    workbook = xlsxwriter.Workbook(QUESTION_PATH)
    worksheet = workbook.add_worksheet()
    worksheet.set_column(1, 1, 250)
    for i, key in enumerate(key_candidates):
        worksheet.write(i, 10, key)
        worksheet.write(i, 1, database[key]['definition'])
    workbook.close()

def write_result():
    workbook = openpyxl.load_workbook(QUESTION_PATH)
    worksheet = workbook.active

    outputs = []
    for row in range(1, worksheet.max_row + 1):
        term = str(worksheet.cell(row=row, column=11).value).strip()
        term_pred = str(worksheet.cell(row=row, column=1).value).strip()
        definition = str(worksheet.cell(row=row, column=2).value).strip()
        outputs.append((term, term_pred, definition))

    workbook = xlsxwriter.Workbook(RESULT_PATH)
    worksheet = workbook.add_worksheet()
    worksheet.set_column(first_col=0, last_col=2, width=50)
    for i, (term, term_pred, definition) in enumerate(outputs):
        worksheet.write(i, 0, term)
        worksheet.write(i, 1, term_pred)
        worksheet.write(i, 2, definition)
    workbook.close()





